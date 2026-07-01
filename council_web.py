# -*- coding: utf-8 -*-
"""
council_web.py — หน้าเว็บ Council Trinity Lab (Flask จิ๋ว + SSE สด)
เปิด: python council_web.py  →  http://127.0.0.1:8091
แยกขาดจากบ้าน (ใช้แค่ core/ ของตัวเอง + flask library)
"""
import os
import sys
import io
import json
import re
import queue
import threading
import hashlib
from datetime import datetime

# กัน console cp874 พังเวลา print emoji/ไทย
try:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
except Exception:
    pass

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from flask import Flask, request, Response, send_from_directory, session
from flask_session import Session

from core import orchestrator, llm, personas, memory, logfilter, applier, router, discussion, cross_memory, code_review, security_review, docgen, error_memory
import glob

# Register cost callback — deduct user credit on every AI call
_current_user = {"username": None}
def _cost_callback(tier, model, usage=None):
    u = _current_user["username"]
    if u:
        _add_user_cost(u, tier, usage)
llm.set_cost_callback(_cost_callback)

_BASE = os.path.dirname(os.path.abspath(__file__))
_CHAT_DIR = os.path.join(_BASE, "chats")
_LOG_DIR = os.path.join(_BASE, "logs")
os.makedirs(_CHAT_DIR, exist_ok=True)
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("FLASK_SECRET_KEY", "council-lab-default-key-please-change")
app.config["SESSION_TYPE"] = "filesystem"
app.config["SESSION_FILE_DIR"] = os.path.join(_BASE, "sessions")
os.makedirs(app.config["SESSION_FILE_DIR"], exist_ok=True)
Session(app)

# entry การประชุมที่เป็น log/ผลรัน (ย่อด้วย logfilter เวลาแนบให้ลูกอ่าน)
_LOGLIKE = ("Auto-Test", "sandbox", "▶️", "🧪", "🛠️")

# User accounts: loaded from users.json (hashed passwords, not in source code)
def _load_users():
    p = os.path.join(_BASE, "users.json")
    if not os.path.exists(p):
        print(f"⚠️  users.json not found at {p} — creating empty")
        return {}
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)

def _hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

_USERS = _load_users()

def _profile_path(username):
    user_dir = os.path.join(_BASE, "chats", username)
    os.makedirs(user_dir, exist_ok=True)
    return os.path.join(user_dir, "_profile.json")

def _has_real_profile(username):
    """ตรวจว่ามี profile จริง (ไม่ใช่แค่กดข้าม)"""
    ppath = _profile_path(username)
    if not os.path.exists(ppath):
        return False
    try:
        with open(ppath, "r", encoding="utf-8") as f:
            pd = json.load(f)
            cg = (pd.get("career_group") or "").strip()
            return bool(cg) and cg != "ข้ามการตั้งค่า"
    except Exception:
        return False

def _load_profile(username):
    """โหลด profile ของ user — คืน dict หรือ {}"""
    ppath = _profile_path(username)
    if not os.path.exists(ppath):
        return {}
    try:
        with open(ppath, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _profile_context_block(username):
    """สร้าง context block จาก profile เพื่อฉีดเข้า AI prompt"""
    pd = _load_profile(username)
    if not pd or not pd.get("career_group") or pd["career_group"] == "ข้ามการตั้งค่า":
        return ""
    parts = []
    parts.append(f"กลุ่มอาชีพ: {pd['career_group']}")
    skills = pd.get("skills") or []
    if skills:
        parts.append(f"ทักษะ: {', '.join(skills)}")
    style = pd.get("style") or []
    if style:
        parts.append(f"สไตล์การทำงาน: {', '.join(style)}")
    proj = pd.get("project") or {}
    if proj.get("name"):
        p_info = proj["name"]
        if proj.get("type"):
            p_info += f" ({proj['type']})"
        if proj.get("path"):
            p_info += f" — {proj['path']}"
        parts.append(f"โปรเจกต์: {p_info}")
    return "\n[โปรไฟล์ผู้ใช้ — ปรับคำตอบให้เข้ากับบริบท]\n" + "\n".join(parts) + "\n"

@app.route("/login", methods=["POST"])
def login():
    data = request.get_json(force=True) or {}
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    
    if username in _USERS and _hash_pw(password) == _USERS[username].get("pw_hash"):
        session["username"] = username
        session["display_name"] = _USERS[username]["name"]
        # ตรวจ profile ฝั่ง server เลย — ส่ง has_profile ไปให้ client
        has_profile = _has_real_profile(username)
        return {"ok": True, "display_name": _USERS[username]["name"], "has_profile": has_profile}
    return {"error": "ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง"}, 401

@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return {"ok": True}


@app.route("/api/topup", methods=["POST"])
def topup():
    """เติมเครดิต — admin เติมให้ user ได้, user เติมให้ตัวเองได้"""
    login_check = require_login()
    if login_check:
        return login_check
    data = request.get_json(force=True) or {}
    amount = float(data.get("amount") or 0)
    target_user = (data.get("username") or session.get("username")).strip()
    if amount <= 0:
        return {"error": "จำนวนเงินไม่ถูกต้อง"}, 400
    if target_user not in _USERS:
        return {"error": "ไม่พบบัญชีผู้ใช้"}, 400
    if session.get("username") != "admin" and target_user != session.get("username"):
        return {"error": "ไม่มีสิทธิ์เติมเงินให้ผู้อื่น"}, 403
    w = _load_user_wallet(target_user)
    w["credit_thb"] = w.get("credit_thb", 0.0) + amount
    w["last_topup"] = datetime.now().isoformat()
    w["last_topup_amount"] = amount
    _save_user_wallet(target_user, w)
    return {"ok": True, "username": target_user, "credit_thb": round(w["credit_thb"], 2)}

def require_login():
    """Check if user is logged in"""
    if not session.get("username"):
        return {"error": "กรุณาเข้าสู่ระบบก่อน"}, 401
    return None


def _list_sessions(limit=15, username=None):
    """รายการประชุมล่าสุดจาก logs/ — ใหม่สุดก่อน
    admin เห็นทั้งหมด, test users เห็นเฉพาะ session ที่ตัวเองรัน"""
    out = []
    for f in sorted(glob.glob(os.path.join(_LOG_DIR, "session_*.json")), reverse=True)[:limit]:
        try:
            d = json.load(open(f, encoding="utf-8"))
        except Exception:
            continue
        # Filter: non-admin users only see their own sessions + legacy (username=None)
        if username and username != "admin":
            su = d.get("username")
            if su is not None and su != username:
                continue
        sid = os.path.basename(f)[8:-5]            # YYYYMMDD_HHMMSS
        label = f"{sid[6:8]}/{sid[4:6]} {sid[9:11]}:{sid[11:13]}"
        out.append({"id": sid, "label": label,
                    "task": (d.get("task") or "")[:60],
                    "task_full": d.get("task") or "",
                    "approved": d.get("approved")})
    return out


def _session_block(sid):
    """ดึงบันทึกการประชุมเต็มของ session นั้นมาเป็นข้อความ (ย่อ log ขยะ) ให้ลูกอ่าน"""
    f = os.path.join(_LOG_DIR, f"session_{sid}.json")
    if not os.path.exists(f):
        return ""
    try:
        d = json.load(open(f, encoding="utf-8"))
    except Exception:
        return ""
    lines = []
    for sp, txt in d.get("transcript", []):
        t = logfilter.compress(txt, 1000) if any(k in sp for k in _LOGLIKE) else (txt or "")
        lines.append(f"[{sp}] {t[:1800]}")
    body = "\n".join(lines)
    if len(body) > 16000:
        body = body[:16000] + "\n…(บันทึกยาว ตัดบางส่วน)…"
    return (f"[บันทึกการประชุมสภาที่คุณร่วมประชุม — งาน: {(d.get('task') or '')[:200]}]\n"
            f"{body}\n[จบบันทึกการประชุม]")


def _session_brief(sid):
    """โหมดประหยัด — ดึงเฉพาะแก่น (โค้ดไฟล์สุดท้าย + ผลเทสล่าสุด + สรุปมติ)
    แทนบทคุยทุกรอบ → input ถูกลง ~50-60% และพี่จูนเห็นโค้ดจริงชัดกว่า diff ในบทสนทนา"""
    f = os.path.join(_LOG_DIR, f"session_{sid}.json")
    if not os.path.exists(f):
        return ""
    try:
        d = json.load(open(f, encoding="utf-8"))
    except Exception:
        return ""
    tr = d.get("transcript", [])

    def last_match(*keys):
        for sp, txt in reversed(tr):
            if any(k in sp for k in keys) and txt:
                return txt
        return ""

    code = last_match("แจงจูน")          # ข้อความ coder ล่าสุด = โค้ดเวอร์ชันสุดท้าย
    test = last_match("Auto-Test")         # ผลเทสล่าสุด
    parts = [f"[งานที่สั่ง]\n{(d.get('task') or '')[:1200]}"]
    if code:
        parts.append(f"[โค้ดไฟล์ล่าสุดที่ส่งมอบ]\n{code[:6000]}")
    if test:
        parts.append(f"[ผล Auto-Test ล่าสุด]\n{logfilter.compress(test, 900)}")
    if d.get("summary"):
        parts.append(f"[สรุปมติของสภา]\n{d['summary'][:2000]}")
    parts.append(f"[ผ่านการตรวจ: {'ใช่' if d.get('approved') else 'ยังไม่ครบ'}]")
    return "[ข้อมูลย่อของการประชุม (โหมดประหยัด)]\n" + "\n\n".join(parts) + "\n[จบข้อมูล]"


def _save_discussion(task, summary, transcript, username=None):
    """บันทึกผลการประชุมราง B (งานคิด) — session log + ส่งมอบข้อสรุปลง projects/
    เหมือนราง A แต่ผลงานคือ structured summary (ไม่มีโค้ดใน workspace)"""
    sid = datetime.now().strftime("%Y%m%d_%H%M%S")
    approved = any("ตกผลึก" in (sp or "") for sp, _ in transcript)

    # 1. บันทึก session log (format เดียวกับราง A ให้ dropdown ขยายความได้)
    try:
        with open(os.path.join(_LOG_DIR, f"session_{sid}.json"), "w", encoding="utf-8") as f:
            json.dump({"task": task, "project": None, "approved": approved,
                       "username": username or "admin",
                       "rounds": len([t for t in transcript if "ฝ่ายค้าน" in (t[0] or "")]),
                       "transcript": transcript, "summary": summary},
                      f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    # 2. ส่งมอบข้อสรุปลง projects/<slug>/summary.md
    slug = orchestrator._slugify(task)
    dest = os.path.join(orchestrator.PROJECTS_DIR, slug)
    if not os.path.realpath(dest).startswith(os.path.realpath(orchestrator.PROJECTS_DIR) + os.sep):
        return ""
    os.makedirs(dest, exist_ok=True)
    try:
        with open(os.path.join(dest, "summary.md"), "w", encoding="utf-8") as f:
            f.write(f"# ข้อสรุปการประชุมสภา (งานคิด)\n\n")
            f.write(f"**โจทย์:** {task[:500]}\n\n")
            f.write(f"**ตกผลึก:** {'ใช่' if approved else 'ยังไม่ตกผลึกเต็มที่'}\n\n---\n\n")
            f.write(summary or "(ไม่มีข้อสรุป)")
    except Exception:
        pass
    return dest


def _chat_path(who, username=None):
    """Generate chat file path - separate per user"""
    def _safe(s):
        return "".join(c for c in (s or "") if c.isalnum() or 0x0e01 <= ord(c) <= 0x0e5b) or "x"
    safe_user = _safe(username or "default")
    safe_who = _safe(who)
    user_dir = os.path.join(_CHAT_DIR, safe_user)
    os.makedirs(user_dir, exist_ok=True)
    return os.path.join(user_dir, f"{safe_who}.json")


import threading as _threading
_chat_locks = {}
_chat_locks_guard = _threading.Lock()

def _get_chat_lock(path):
    with _chat_locks_guard:
        if path not in _chat_locks:
            _chat_locks[path] = _threading.Lock()
        return _chat_locks[path]

def _load_chat(who, username=None):
    path = _chat_path(who, username)
    lock = _get_chat_lock(path)
    with lock:
        try:
            with open(path, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []


def _save_chat(who, msgs, username=None):
    path = _chat_path(who, username)
    lock = _get_chat_lock(path)
    with lock:
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(msgs[-200:], f, ensure_ascii=False, indent=2)
        except Exception:
            pass

_events = queue.Queue()
_running = {"on": False}
_running_id = {"id": None}          # track current run request ID
_stop_flags = {}                    # per-request stop flags: {req_id: bool}
_audit_answers = {}                 # per-request audit answers: {req_id: "all"/"high"/"skip"}
import uuid as _uuid


@app.after_request
def _cors(resp):
    # อนุญาตเฉพาะ origin ที่รู้จัก — ไม่เปิด * ป้องกัน CSRF
    origin = request.headers.get("Origin", "")
    allowed = ("http://127.0.0.1:8091", "http://localhost:8091",
               "http://192.168.40.158:8091", "https://council.inthasapdigital.com")
    if origin and any(origin.startswith(a) for a in allowed):
        resp.headers["Access-Control-Allow-Origin"] = origin
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return resp


# ── Rate limiting: กัน spam — สูงสุด 30 req/min ต่อ user ──
_rate_limits = {}  # {username: [(timestamp, ...), ...]}
_rate_window = 60  # วินาที
_rate_max = 30     # max requests per window

def _check_rate_limit(username):
    """คืน True ถ้าผ่าน, False ถ้าเกิน limit"""
    import time as _time
    now = _time.time()
    if username not in _rate_limits:
        _rate_limits[username] = []
    # กรองเฉพาะที่อยู่ใน window
    _rate_limits[username] = [t for t in _rate_limits[username] if now - t < _rate_window]
    if len(_rate_limits[username]) >= _rate_max:
        return False
    _rate_limits[username].append(now)
    return True


@app.route("/api/health")
def health_check():
    """Health check endpoint — สำหรับ auto-restart/monitoring"""
    try:
        tiers = llm.available_tiers()
        return {
            "status": "ok",
            "tiers": tiers,
            "running": _running["on"],
            "uptime": datetime.now().isoformat()
        }
    except Exception as e:
        return {"status": "error", "message": str(e)[:200]}, 500


@app.route("/")
@app.route("/index.html")
def index():
    resp=send_from_directory(os.path.join(_BASE, "web"), "index.html")
    resp.headers["Cache-Control"]="no-cache, no-store, must-revalidate"
    resp.headers["Pragma"]="no-cache"
    return resp

@app.route("/manual.html")
def manual():
    return send_from_directory(os.path.join(_BASE, "web"), "manual.html")

@app.route("/profile.html")
def profile_page():
    resp = send_from_directory(os.path.join(_BASE, "web"), "profile.html")
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp

@app.route("/research.html")
def research():
    return send_from_directory(os.path.join(_BASE, "web"), "research.html")


@app.route("/api/tiers")
def tiers():
    login_check = require_login()
    if login_check:
        return login_check
    return {"tiers": llm.available_tiers(), "budget_today": round(llm._today_cost(), 4),
            "credit_thb": round(_get_user_credit_thb(session.get("username")), 2)}


@app.route("/api/run", methods=["POST"])
def run():
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    if not _has_credit(username):
        return {"error": "เครดิตหมดแล้ว กรุณาเติมเงิน 50 บาท เพื่อใช้งานต่อ\nธนาคารกสิกรไทย เลขบัญชี 262-2-80739-7"}, 402
    if _running["on"]:
        return {"error": "สภากำลังประชุมอยู่ รอรอบนี้จบก่อน"}, 409
    data = request.get_json(force=True) or {}
    task = (data.get("task") or "").strip()
    project = (data.get("project") or "").strip() or None
    apply_to = project if data.get("apply") else None     # แก้ไฟล์จริงเฉพาะเมื่อพ่อติ๊ก
    try:
        rounds = max(1, min(6, int(data.get("rounds") or 4)))   # เพดาน 6 (ปลอดภัยเพราะมีเบรกหยุดอัตโนมัติ) · default 4
    except (TypeError, ValueError):
        rounds = 3
    if not task:
        return {"error": "ยังไม่ได้พิมพ์งาน"}, 400
    if apply_to and not project:
        return {"error": "ติ๊ก 'แก้ไฟล์จริง' แล้วต้องใส่ path โปรเจกต์ในช่องโปรเจกต์ด้วย"}, 400

    while not _events.empty():
        _events.get_nowait()

    req_id = str(_uuid.uuid4())
    _stop_flags[req_id] = False
    _running_id["id"] = req_id
    _current_user["username"] = username
    mode = (data.get("mode") or "auto").strip().lower()   # auto / a (งานโค้ด) / b (งานคิด)

    def _emit(s, t):
        _events.put({"speaker": s, "text": t})

    def worker():
        _running["on"] = True
        try:
            chosen = mode
            if mode == "auto":                            # Router จำแนกให้
                r = router.classify(task)
                _emit("🧭 จำแนกงาน",
                      f"→ ราง {r['route']} · มั่นใจ {r['confidence']:.0%} · {r['action']}\n{r['reason']}")
                if r["route"] == "B":
                    chosen = "b"
                elif r["route"] == "MIXED":
                    _emit("🧭 หมายเหตุ", "งานนี้ปนทั้งคิดและโค้ด — ทำราง A (โค้ด) ก่อน "
                          "ถ้าอยากให้สภาถกเชิงความคิด เลือกโหมด 'งานคิด' เองได้")
                    chosen = "a"
                else:
                    chosen = "a"
            if chosen == "b":                             # ราง B — สภาถกงานคิด
                d_summary, d_transcript = discussion.run_discussion(
                    task, max_rounds=rounds,
                    on_event=_emit, should_stop=lambda: _stop_flags.get(req_id, False))
                # บันทึกผลการประชุมราง B (session log + ส่งมอบข้อสรุปลง projects/)
                try:
                    dest = _save_discussion(task, d_summary, d_transcript, username=username)
                    if dest:
                        rel = os.path.relpath(dest, _BASE)
                        crystallized = any("ตกผลึก" in (sp or "") for sp, _ in d_transcript)
                        _emit("📦 ส่งมอบงาน",
                              f"บันทึกข้อสรุปไว้ที่: {rel}\\summary.md\n"
                              f"(เปิดอ่านได้เลย){'  ✅ ตกผลึก' if crystallized else '  ⚠️ ยังไม่ตกผลึกเต็มที่'}")
                except Exception as e:
                    _emit("📦 ส่งมอบงาน", f"(บันทึกไม่สำเร็จ: {str(e)[:120]})")
            else:                                         # ราง A — เขียน/แก้โค้ด (เดิม)
                orchestrator.run_task(
                    task, project_path=project, max_rounds=rounds,
                    on_event=_emit, should_stop=lambda: _stop_flags.get(req_id, False), apply_to=apply_to,
                    username=username)
        except Exception as e:
            _emit("⚠️ error", str(e)[:300])
        finally:
            # บันทึกความจำข้ามเซสชั่น — สภาทำงานอะไรเสร็จ
            try:
                cross_memory.add_fact(username, f"ส่งงานให้สภา: {task[:100]}", "project")
            except Exception:
                pass
            _emit("__done__", "")
            _running["on"] = False
            _stop_flags.pop(req_id, None)
            _running_id["id"] = None

    threading.Thread(target=worker, daemon=True).start()
    return {"ok": True}


@app.route("/api/stop", methods=["POST"])
def stop():
    login_check = require_login()
    if login_check:
        return login_check
    rid = _running_id.get("id")
    if rid and rid in _stop_flags:
        _stop_flags[rid] = True
    return {"ok": True}


@app.route("/api/pick-folder")
def pick_folder():
    """เปิด native OS folder picker (tkinter) — ข้าม browser security ได้เพราะรันฝั่ง server
    ใช้ได้เฉพาะเครื่อง local (localhost) เท่านั้น"""
    login_check = require_login()
    if login_check:
        return login_check
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        folder = filedialog.askdirectory(title="เลือกโฟลเดอร์โปรเจกต์")
        root.destroy()
        if folder:
            return {"path": folder}
        return {"path": ""}
    except Exception as e:
        return {"path": "", "error": str(e)[:200]}


@app.route("/api/audit", methods=["POST"])
def audit():
    """ตรวจสอบบั๊กทั้งโปรเจกต์ — สแกน → แก้ → วนจนเสร็จ"""
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    if not _has_credit(username):
        return {"error": "เครดิตหมดแล้ว กรุณาเติมเงิน 50 บาท เพื่อใช้งานต่อ\nธนาคารกสิกรไทย เลขบัญชี 262-2-80739-7"}, 402
    if _running["on"]:
        return {"error": "สภากำลังทำงานอยู่ รอรอบนี้จบก่อน"}, 409
    data = request.get_json(force=True) or {}
    project = (data.get("project") or "").strip()
    if not project:
        return {"error": "ต้องระบุ path โปรเจกต์ที่ต้องการตรวจสอบ"}, 400

    while not _events.empty():
        _events.get_nowait()

    req_id = str(_uuid.uuid4())
    _stop_flags[req_id] = False
    _running_id["id"] = req_id
    _current_user["username"] = username
    _audit_answers[req_id] = None

    def _emit(s, t):
        _events.put({"speaker": s, "text": t})

    def _get_answer(timeout=120):
        import time as _time
        deadline = _time.time() + timeout
        while _time.time() < deadline:
            if _stop_flags.get(req_id, False):
                return None
            if _audit_answers.get(req_id) is not None:
                return _audit_answers.pop(req_id)
            _time.sleep(0.5)
        return None

    def worker():
        _running["on"] = True
        try:
            orchestrator.run_project_audit(
                project,
                on_event=_emit,
                should_stop=lambda: _stop_flags.get(req_id, False),
                username=username,
                get_answer=_get_answer)
        except Exception as e:
            _emit("⚠️ error", str(e)[:300])
        finally:
            try:
                cross_memory.add_fact(username, f"ตรวจสอบบั๊กโปรเจกต์: {project[:100]}", "project")
            except Exception:
                pass
            _emit("__done__", "")
            _running["on"] = False
            _stop_flags.pop(req_id, None)
            _audit_answers.pop(req_id, None)
            _running_id["id"] = None

    threading.Thread(target=worker, daemon=True).start()
    return {"ok": True}


@app.route("/api/audit/answer", methods=["POST"])
def audit_answer():
    """รับคำตอบจากผู้ใช้สำหรับคำถามระหว่าง audit"""
    login_check = require_login()
    if login_check:
        return login_check
    data = request.get_json(force=True) or {}
    answer = (data.get("answer") or "").strip()
    if not answer:
        return {"error": "ไม่มีคำตอบ"}, 400
    rid = _running_id.get("id")
    if rid and rid in _audit_answers:
        _audit_answers[rid] = answer
        return {"ok": True}
    return {"error": "ไม่มีคำถามรอคำตอบ"}, 404


# ── โหมดคุยรายบุคคล (ตัวต่อตัว ไม่ใช่ประชุมสภา) ──
def _build_chat_wrap(username):
    """สร้าง chat wrap ตามบริบทผู้ใช้ — admin เรียก 'อาจารย์', test user เรียก 'คุณ'
    ถ้ามี profile ใช้ display_name จาก profile"""
    title = "อาจารย์" if username == "admin" else "คุณ"
    # ใช้ชื่อจาก profile ถ้ามี
    pd = _load_profile(username)
    if pd and pd.get("display_name"):
        title = pd["display_name"]
    return (
        f"\n\n[หมายเหตุ: นี่คือการให้คำปรึกษาแบบตัวต่อตัว ไม่ใช่ประชุมสภา และไม่ใช่โหมดเขียนโค้ด]\n"
        f"‼️ กฎสำคัญ: ตอบตรงคำถามก่อนเสมอ — ใช้ศัพท์วิชาการเมื่อเหมาะสม อย่าใช้ความรู้สึกมาแทนคำตอบ\n"
        f"พูดอย่างมืออาชีพ ตรงประเด็น เด็ดขาด น่าเชื่อถือ\n"
        f"เรียกผู้ใช้ว่า '{title}' แทนตัวเองด้วยชื่อของคุณเอง "
        f"ลงท้ายประโยคด้วย 'ค่ะ' ห้ามขี้อ้อน ห้าม 'เจ้าค่ะ' ห้าม 'ผม/ครับ'\n"
        f"ในบทสนทนา: [{title}]={title}พูด · [พี่จูน]=พี่จูน (Claude) พูด · "
        f"ชื่ออื่น=สมาชิกคนอื่นในทีม\n"
        f"‼️ เกี่ยวกับโค้ด: เขียนโค้ดให้เฉพาะเมื่อ{title}ขอให้เขียนโค้ดหรือแสดงโค้ดโดยตรงเท่านั้น "
        f"ถ้า{title}ไม่ได้กล่าวถึงโค้ดเลย ห้ามเขียนโค้ด ห้ามใช้ code block ห้ามใช้ token RUN: หรือ file= "
        f"(โหมดนี้ไม่ได้รันใน sandbox — ถ้า{title}ขอโค้ดจริง ให้เขียน code block ธรรมดาโดยไม่ใช้ file=)\n"
        f"ถ้า{title}ถามเรื่องทั่วไป/งาน/ประชาสัมพันธ์/เอกสาร ให้ตอบเป็นข้อความปกติเท่านั้น"
    )


def _chat_prompt(msgs, msg, title="คุณ"):
    """สร้าง prompt แชทเดี่ยว — ไม่ตัดจดหมายทิ้ง (ย่อเฉพาะเมื่อยาวจริงๆ)"""
    def label(r):
        if r in ("พ่อ", title):  # รองรับ chat เก่าที่เก็บ role="พ่อ"
            return title
        if r == "พี่จูน":
            return "พี่จูน(Claude)"
        return r
    lines = [f"[{label(m.get('role'))}] {m.get('text') or ''}" for m in msgs if m.get("text")]
    hist = "\n".join(lines)
    if len(hist) > 7000:     # ยาวมากค่อยย่อด้วย rolling summary
        hist = memory.build_context([(label(m.get("role")), m.get("text") or "") for m in msgs])
    head = f"บทสนทนาก่อนหน้าในห้องนี้:\n{hist}\n\n" if hist else ""
    return head + f"{title}เพิ่งพูดว่า: {msg}"


@app.route("/api/chat", methods=["POST"])
def chat():
    # Check login
    login_check = require_login()
    if login_check:
        return login_check
    
    username = session.get("username")
    if not _has_credit(username):
        return {"error": "เครดิตหมดแล้ว กรุณาเติมเงิน 50 บาท เพื่อใช้งานต่อ\nธนาคารกสิกรไทย เลขบัญชี 262-2-80739-7"}, 402
    if not _check_rate_limit(username):
        return {"error": "ส่งเร็วเกินไป — รอสักครู่แล้วลองใหม่ (จำกัด 30 ครั้ง/นาที)"}, 429
    data = request.get_json(force=True) or {}
    who = (data.get("persona") or "").strip()
    msg = (data.get("message") or "").strip()
    history = data.get("history") or []      # [{role, text}, ...]
    if who not in personas.PERSONAS:
        return {"error": "ไม่รู้จักตัวแทนนี้"}, 400
    if not msg:
        return {"error": "ยังไม่ได้พิมพ์ข้อความ"}, 400

    p = personas.PERSONAS[who]
    username = session.get("username")
    msgs = _load_chat(who, username)          # ประวัติถาวรจากดิสก์ (รวมจดหมายพี่จูนถ้ามี)
    title = "อาจารย์" if username == "admin" else "คุณ"
    # ใช้ชื่อจาก profile ถ้ามี
    pd = _load_profile(username)
    if pd and pd.get("display_name"):
        title = pd["display_name"]

    def _sanitize_path(raw):
        """Sanitize user-provided path — block .. and dangerous patterns"""
        if not raw:
            return ""
        raw = raw.strip()
        if ".." in raw or "\x00" in raw:
            return ""
        return raw

    # ── ปั๊กอิน: คำสั่งพิเศษในแชท — /review /security /plan ──
    msg_lower = msg.strip().lower()
    _current_user["username"] = username
    if msg_lower.startswith("/review"):
        # /review [path] — รีวิว 5 เอเจนต์
        parts = msg.split(None, 1)
        rp = _sanitize_path(parts[1]) if len(parts) > 1 else ""
        if rp == "" and len(parts) > 1:
            return {"reply": "⚠️ path ต้องไม่มี .. หรืออักขระอันตราย", "tier": "tool", "persona": who}
        msgs.append({"role": title, "text": msg})
        msgs.append({"role": who, "text": "⏳ กำลังรีวิวโค้ด 5 มุมมอง รอสักครู่...", "tier": "tool"})
        _save_chat(who, msgs, username)
        try:
            _chat_stop = {"flag": False}
            rv_summary, rv_findings = code_review.run_review(rp or _BASE, should_stop=lambda: _chat_stop["flag"])
            tool_reply = f"🔍 รีวิว 5 เอเจนต์\n{rv_summary}\n\n"
            for key, f in rv_findings.items():
                tool_reply += f"\n{f['name']} ({f.get('tier','')}):\n{f.get('report','')[:1500]}\n"
        except Exception as e:
            tool_reply = f"รีวิวไม่สำเร็จ: {e}"
        msgs.append({"role": who, "text": tool_reply, "tier": "tool"})
        _save_chat(who, msgs, username)
        cross_memory.add_fact(username, f"ใช้ /review ในแชท: {msg[:60]}", "note")
        return {"reply": tool_reply, "tier": "tool", "persona": who}

    if msg_lower.startswith("/security"):
        # /security [path] — สแกนช่องโหว่
        parts = msg.split(None, 1)
        sp = _sanitize_path(parts[1]) if len(parts) > 1 else ""
        if sp == "" and len(parts) > 1:
            return {"reply": "⚠️ path ต้องไม่มี .. หรืออักขระอันตราย", "tier": "tool", "persona": who}
        msgs.append({"role": title, "text": msg})
        msgs.append({"role": who, "text": "⏳ กำลังสแกนช่องโหว่ รอสักครู่...", "tier": "tool"})
        _save_chat(who, msgs, username)
        try:
            _chat_stop = {"flag": False}
            sec_summary, sec_findings, sec_ai = security_review.run_security_review(
                sp or _BASE, should_stop=lambda: _chat_stop["flag"])
            tool_reply = f"🔒 ซีเคียวริตี้ รีวิว\n{sec_summary}\n\n"
            sf = sec_findings[:15]
            for f in sf:
                tool_reply += f"  [{f['severity']}] {f['file']}:{f['line']} — {f['description']}\n"
            if sec_ai:
                tool_reply += f"\n🤖 AI Audit:\n{sec_ai[:1500]}\n"
        except Exception as e:
            tool_reply = f"สแกนไม่สำเร็จ: {e}"
        msgs.append({"role": who, "text": tool_reply, "tier": "tool"})
        _save_chat(who, msgs, username)
        cross_memory.add_fact(username, f"ใช้ /security ในแชท: {msg[:60]}", "note")
        return {"reply": tool_reply, "tier": "tool", "persona": who}

    if msg_lower.startswith("/plan"):
        # /plan <task> — วางแผน + test strategy
        parts = msg.split(None, 1)
        plan_task = parts[1].strip() if len(parts) > 1 else ""
        if not plan_task:
            return {"reply": "ใช้: /plan <งานที่ต้องการวางแผน>", "tier": "tool", "persona": who}
        msgs.append({"role": title, "text": msg})
        msgs.append({"role": who, "text": "⏳ กำลังวางแผน รอสักครู่...", "tier": "tool"})
        _save_chat(who, msgs, username)
        try:
            plan_prompt = (
                f"งาน:\n{plan_task}\n\n"
                "สร้าง 'แผนปฏิบัติ' ละเอียดระดับไฟล์ — มี:\n"
                "1. ไฟล์ที่ต้องสร้าง/แก้ (ชื่อไฟล์ + หน้าที่สั้นๆ)\n"
                "2. ลำดับการทำ (ขั้นตอน 1→2→3)\n"
                "3. 'Test Strategy': test case ที่ต้องเขียน พร้อม input → expected output\n"
                "   ระบุ test case ขั้นต่ำ 3 กรณี: (ก) กรณีปกติ (ข) กรณี edge case (ค) กรณี error\n"
                "4. ความเสี่ยงที่อาจทำให้ติด และวิธีกัน\n"
                "ตอบสั้น กระชับ อ่านได้ใน 1 นาที"
            )
            mem_block = cross_memory.context_block(username)
            if mem_block:
                plan_prompt = mem_block + plan_prompt
            plan_reply, pt, _ = llm.call_tier("deepseek", p["system"], plan_prompt, max_tokens=3000)
            tool_reply = f"📋 แผนปฏิบัติ [{pt}]\n\n{plan_reply}"
        except Exception as e:
            pt = "tool"
            tool_reply = f"วางแผนไม่สำเร็จ: {e}"
        msgs.append({"role": who, "text": tool_reply, "tier": pt})
        _save_chat(who, msgs, username)
        cross_memory.add_fact(username, f"ใช้ /plan ในแชท: {plan_task[:60]}", "project")
        return {"reply": tool_reply, "tier": pt, "persona": who}

    if msg_lower.startswith("/memory"):
        # /memory — ดูความจำสะสม
        mem = cross_memory.load_memory(username)
        facts = mem.get("facts", [])
        if not facts:
            tool_reply = "🧠 ยังไม่มีความจำสะสม — AI จะเริ่มจำเมื่อคุณใช้งาน"
        else:
            tool_reply = f"🧠 ความจำสะสม ({len(facts)} รายการ):\n"
            for f in reversed(facts):
                icon = "📦" if f.get("type") == "project" else "💬"
                tool_reply += f"  {icon} {f.get('text','')} ({f.get('time','')})\n"
        msgs.append({"role": title, "text": msg})
        msgs.append({"role": who, "text": tool_reply, "tier": "tool"})
        _save_chat(who, msgs, username)
        return {"reply": tool_reply, "tier": "tool", "persona": who}

    prompt = _chat_prompt(msgs, msg, title=title)
    # ฉีดโปรไฟล์ผู้ใช้ — ให้ AI รู้บริบทอาชีพ/ทักษะ/สไตล์
    profile_block = _profile_context_block(username)
    if profile_block:
        prompt = profile_block + prompt
    # ฉีดความจำข้ามเซสชั่น — ให้ AI รู้ว่า user ทำอะไรมาก่อน
    mem_block = cross_memory.context_block(username)
    if mem_block:
        prompt = mem_block + prompt
    sid = (data.get("session_id") or "").strip()
    if sid:                                      # แนบบันทึกการประชุมที่พ่อเลือกให้ลูกอ่าน
        blk = _session_block(sid)
        if blk:
            prompt = blk + "\n\n" + prompt

    # รับรูปแนบ — ถ้ามี ต้องสลับไป vision-capable tier (Gemini/Claude)
    images = data.get("images") or []
    preferred = p["model"]
    vision_note = ""
    if images:
        if preferred == "deepseek":
            # DeepSeek ไม่รองรับ vision → สลับไป Gemini เท่านั้น (Claude แพงเกินไป)
            if llm._key_for("gemini"):
                preferred = "gemini"
                vision_note = f"(สลับจาก {p['model']} → gemini เพราะมีรูปแนบ)"
            else:
                return {"error": "โมเดลที่ใช้อยู่ (deepseek) ไม่รองรับรูปภาพ และไม่มี Gemini key ในระบบ"}, 400

    _current_user["username"] = username
    try:
        reply, tier, _ = llm.call_tier(preferred, p["system"] + _build_chat_wrap(username), prompt,
                                       max_tokens=6000, images=images if images else None)
    except Exception as e:
        return {"error": str(e)[:300]}, 500
    msgs.append({"role": title, "text": msg})
    reply_text = reply + (f"\n\n_{vision_note}_" if vision_note else "")
    msgs.append({"role": who, "text": reply_text, "tier": tier})
    _save_chat(who, msgs, username)
    # บันทึกความจำข้ามเซสชั่น — คุยเรื่องอะไรกับใคร
    try:
        cross_memory.add_fact(username, f"คุยกับ {who}: {msg[:80]}", "note")
    except Exception:
        pass
    return {"reply": reply_text, "tier": tier, "persona": who}


@app.route("/api/genimage", methods=["POST"])
def genimage():
    """สร้างรูปด้วย Pollinations.ai (ฟรี ไม่ต้อง API key)
    Anonymous: 1 คำขอ / 15 วินาที, รูปมีลายน้ำ
    สมัครฟรี: 1 คำขอ / 5 วินาที, ไม่มีลายน้ำ"""
    login_check = require_login()
    if login_check:
        return login_check
    data = request.get_json(force=True) or {}
    prompt = (data.get("prompt") or "").strip()
    if not prompt:
        return {"error": "กรุณาพิมพ์คำอธิบายรูปที่ต้องการสร้าง"}, 400
    if len(prompt) > 500:
        return {"error": "คำอธิบายยาวเกินไป (สูงสุด 500 ตัวอักษร)"}, 400

    import urllib.parse, urllib.request, base64
    encoded = urllib.parse.quote(prompt)
    url = f"https://image.pollinations.ai/prompt/{encoded}?width=512&height=512&nologo=true&model=flux"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "CouncilTrinityLab/1.0"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            img_data = resp.read()
        if len(img_data) < 1000:
            return {"error": "ไม่สามารถสร้างรูปได้ กรุณาลองใหม่"}, 502
        b64 = base64.b64encode(img_data).decode("ascii")
        mime = resp.headers.get("Content-Type", "image/jpeg")
        return {"image": f"data:{mime};base64,{b64}", "prompt": prompt,
                "limit_info": "Pollinations.ai — ฟรี ไม่จำกัดจำนวนรูป แต่คำขอติดๆ กันต้องห่าง 15 วินาที (Anonymous) หรือ 5 วินาที (สมัครฟรี)"}
    except urllib.error.HTTPError as e:
        return {"error": f"Pollinations.ai ปฏิเสธคำขอ (HTTP {e.code}) — อาจถูก rate limit กรุณารอ 15 วินาทีแล้วลองใหม่"}, 502
    except Exception as e:
        return {"error": f"ไม่สามารถสร้างรูปได้: {str(e)[:200]}"}, 500


# ── สแกนบิล → Excel ──
_easyocr_reader = None
def _get_easyocr():
    global _easyocr_reader
    if _easyocr_reader is None:
        import easyocr
        _easyocr_reader = easyocr.Reader(["th", "en"], gpu=False)
    return _easyocr_reader

@app.route("/api/scan_bill", methods=["POST"])
def scan_bill():
    """สแกนบิล: EasyOCR ก่อน → ถ้า confidence ต่ำ ส่ง Gemini Vision"""
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    if not _has_credit(username):
        return {"error": "เครดิตหมดแล้ว กรุณาเติมเงิน 50 บาท เพื่อใช้งานต่อ\nธนาคารกสิกรไทย เลขบัญชี 262-2-80739-7"}, 402

    data = request.get_json(force=True) or {}
    image_b64 = (data.get("image") or "").strip()
    if not image_b64:
        return {"error": "กรุณาอัปโหลดรูปบิล"}, 400

    # ถอด base64
    import base64, io
    try:
        header, b64data = image_b64.split(",", 1) if "," in image_b64 else ("", image_b64)
        img_bytes = base64.b64decode(b64data)
    except Exception:
        return {"error": "รูปแบบไฟล์ไม่ถูกต้อง"}, 400

    # ── ชั้น 1: EasyOCR (ฟรี) ──
    ocr_text = ""
    ocr_confidence = 0.0
    try:
        import numpy as np
        from PIL import Image
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
        reader = _get_easyocr()
        results = reader.readtext(np.array(img))
        if results:
            texts = []
            confs = []
            for (bbox, text, conf) in results:
                texts.append(text)
                confs.append(conf)
            ocr_text = "\n".join(texts)
            ocr_confidence = sum(confs) / len(confs) if confs else 0.0
    except Exception:
        ocr_text = ""
        ocr_confidence = 0.0

    # ── ชั้น 2: Gemini สกัดข้อมูลเป็น JSON (เสมอ) ──
    # EasyOCR สกัด raw text ได้ → ส่ง text ให้ Gemini parse เป็น JSON (ถูก, ใช้ text model)
    # EasyOCR ไม่ได้ผล → ส่งรูปให้ Gemini Vision (แพงกว่า)
    used_vision = False
    if llm._key_for("gemini"):
        try:
            _current_user["username"] = username
            if ocr_text.strip():
                # มี raw text จาก EasyOCR แล้ว ให้ Gemini parse เป็น JSON (ประหยัด token)
                prompt = (
                    "จากข้อความ OCR ต่อไปนี้ สกัดข้อมูลทั้งหมดที่เจอในบิล/ใบเสร็จ/statement "
                    "เป็น JSON แบบ flat key-value (ไม่มี nested object ไม่มี array) "
                    "ใช้ key เป็นชื่อหัวข้อในภาษาไทย เช่น วันที่, เวลา, รหัสอ้างอิง, จาก, ไปยัง, "
                    "บัญชีต้นทาง, บัญชีปลายทาง, จำนวนเงิน, ธนาคาร, สถานะ, ร้าน/ผู้ขาย, เลขที่บิล, "
                    "รายการ, รวมทั้งหมด, ภาษี, ประเภท, ค่าธรรมเนียม, รหัสร้านค้า, รหัสธุรกรรม "
                    "สกัดทุกข้อมูลที่เห็น ถ้ามีหลายรายการให้ใส่ใน key 'รายการ' เป็น string รวมทั้งหมด "
                    "ตอบเฉพาะ JSON เท่านั้น ไม่ต้องมีคำอธิบาย ไม่ต้องมี markdown code fence\n\n"
                    "ข้อความ OCR:\n" + ocr_text
                )
                reply, tier, _ = llm.call_tier("gemini", "", prompt, max_tokens=2000)
            else:
                # ไม่มี text เลย ส่งรูปให้ Gemini Vision
                prompt = (
                    "สกัดข้อมูลทั้งหมดที่เห็นในบิล/ใบเสร็จ/statement นี้ "
                    "เป็น JSON แบบ flat key-value (ไม่มี nested object ไม่มี array) "
                    "ใช้ key เป็นชื่อหัวข้อในภาษาไทย เช่น วันที่, เวลา, รหัสอ้างอิง, จาก, ไปยัง, "
                    "บัญชีต้นทาง, บัญชีปลายทาง, จำนวนเงิน, ธนาคาร, สถานะ, ร้าน/ผู้ขาย, เลขที่บิล, "
                    "รายการ, รวมทั้งหมด, ภาษี, ประเภท, ค่าธรรมเนียม, รหัสร้านค้า, รหัสธุรกรรม "
                    "สกัดทุกข้อมูลที่เห็น ถ้ามีหลายรายการให้ใส่ใน key 'รายการ' เป็น string รวมทั้งหมด "
                    "ตอบเฉพาะ JSON เท่านั้น ไม่ต้องมีคำอธิบาย ไม่ต้องมี markdown code fence"
                )
                img_data = b64data if b64data else image_b64
                reply, tier, _ = llm.call_tier("gemini", "", prompt, max_tokens=2000,
                                               images=[{"mime": "image/jpeg", "data": img_data}])
                used_vision = True
            if reply and reply.strip():
                ocr_text = reply.strip()
                ocr_confidence = 0.9
        except Exception:
            pass

    if not ocr_text.strip():
        return {"error": "ไม่สามารถสกัดข้อมูลจากรูปได้ กรุณาถ่ายรูปใหม่ให้ชัดขึ้น"}, 400

    return {"text": ocr_text, "confidence": round(ocr_confidence, 2),
            "method": "gemini_vision" if used_vision else "easyocr"}


@app.route("/api/export_excel", methods=["POST"])
def export_excel():
    """รับ JSON rows → สร้าง .xlsx → ส่งกลับไฟล์"""
    login_check = require_login()
    if login_check:
        return login_check
    data = request.get_json(force=True) or {}
    rows = data.get("rows") or []
    filename = (data.get("filename") or "bill_export").strip()
    if not rows:
        return {"error": "ไม่มีข้อมูลส่งออก"}, 400

    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "บิล"

    # หัวตาราง
    headers = list(rows[0].keys())
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # ข้อมูล
    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(headers, 1):
            val = row.get(h, "")
            # ── Data Sanitization Layer ──
            # ล้างค่าทุกแบบก่อนเข้า Excel: ลบ "บาท", "฿", comma, space, แล้วแปลงเป็น float
            if isinstance(val, str):
                cleaned = val.strip()
                # ลบหน่วยเงินและเครื่องหมายแปลกปลอม
                cleaned = re.sub(r'[บาท฿฿]', '', cleaned)
                cleaned = cleaned.replace(',', '')
                cleaned = cleaned.strip()
                # ถ้าเหลือแค่ตัวเลข + ทศนิยม → แปลงเป็น float
                if re.match(r'^-?\d+\.?\d*$', cleaned):
                    try:
                        val = float(cleaned)
                    except ValueError:
                        pass
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left" if isinstance(val, str) else "right",
                                       vertical="center", wrap_text=True)

    # ปรับความกว้างคอลัมน์
    for col_idx, h in enumerate(headers, 1):
        max_len = max(len(str(h)), max((len(str(row.get(h, ""))) for row in rows), default=10))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)

    # แถวผลรวม (ถ้ามีคอลัมน์ amount/จำนวนเงิน/total)
    amount_cols = [i for i, h in enumerate(headers) if any(k in h.lower() for k in ["amount", "จำนวนเงิน", "total", "รวม"])]
    if amount_cols:
        sum_row = r_idx + 1
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=sum_row, column=c_idx)
            cell.border = thin_border
            cell.font = Font(bold=True)
            if c_idx - 1 in amount_cols:
                col_letter = ws.cell(row=1, column=c_idx).column_letter
                cell.value = f"=SUM({col_letter}2:{col_letter}{r_idx})"
                cell.alignment = Alignment(horizontal="right")
            elif c_idx == 1:
                cell.value = "รวมทั้งหมด"

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_bytes = buf.read()

    # เก็บไฟล์ไว้ที่ server ด้วย
    import os as _os, time as _time
    export_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "exports")
    _os.makedirs(export_dir, exist_ok=True)
    safe_filename = filename.replace("/", "_").replace("\\", "_") + ".xlsx"
    filepath = _os.path.join(export_dir, safe_filename)
    with open(filepath, "wb") as f:
        f.write(file_bytes)

    import urllib.parse
    safe_name = urllib.parse.quote(filename)
    return Response(file_bytes, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename*=UTF-8''{safe_name}.xlsx"})


@app.route("/api/list_exports")
def list_exports():
    login_check = require_login()
    if login_check:
        return login_check
    import os as _os, time as _time
    export_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "exports")
    if not _os.path.isdir(export_dir):
        return {"files": []}
    files = []
    for fn in sorted(_os.listdir(export_dir), key=lambda x: _os.path.getmtime(_os.path.join(export_dir, x)), reverse=True):
        if fn.endswith(".xlsx"):
            mtime = _os.path.getmtime(_os.path.join(export_dir, fn))
            files.append({"name": fn, "size": _os.path.getsize(_os.path.join(export_dir, fn)),
                          "time": _time.strftime("%d/%m/%Y %H:%M", _time.localtime(mtime))})
    return {"files": files}


@app.route("/api/download_export/<path:filename>")
def download_export(filename):
    login_check = require_login()
    if login_check:
        return login_check
    import os as _os
    export_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "exports")
    return send_from_directory(export_dir, filename, as_attachment=True)


@app.route("/api/memory")
def get_memory():
    """ดูความจำข้ามเซสชั่นของ user"""
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    return cross_memory.load_memory(username)


@app.route("/api/memory/clear", methods=["POST"])
def clear_memory():
    """ล้างความจำข้ามเซสชั่น"""
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    cross_memory.save_memory(username, {"facts": []})
    return {"ok": True}


@app.route("/api/code_review", methods=["POST"])
def code_review_endpoint():
    """รีวิวโค้ด 5 เอเจนต์ — รับ project_path ส่งกลับผลรีวิว"""
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    if not _has_credit(username):
        return {"error": "เครดิตหมดแล้ว กรุณาเติมเงิน 50 บาท เพื่อใช้งานต่อ\nธนาคารกสิกรไทย เลขบัญชี 262-2-80739-7"}, 402
    data = request.get_json(force=True) or {}
    project_path = (data.get("project_path") or "").strip()
    if not project_path:
        project_path = _BASE  # ถ้าไม่ระบุ ใช้โปรเจกต์ตัวเอง
    if not os.path.isdir(project_path):
        return {"error": f"ไม่พบโฟลเดอร์: {project_path}"}, 400

    _current_user["username"] = username
    _api_stop = {"flag": False}
    summary, findings = code_review.run_review(
        project_path,
        should_stop=lambda: _api_stop["flag"]
    )
    return {"summary": summary, "findings": findings}


@app.route("/api/security_review", methods=["POST"])
def security_review_endpoint():
    """ซีเคียวริตี้ รีวิว — สแกน codebase หาช่องโหว่ก่อนขึ้น live"""
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    if not _has_credit(username):
        return {"error": "เครดิตหมดแล้ว กรุณาเติมเงิน 50 บาท เพื่อใช้งานต่อ\nธนาคารกสิกรไทย เลขบัญชี 262-2-80739-7"}, 402
    data = request.get_json(force=True) or {}
    project_path = (data.get("project_path") or "").strip()
    if not project_path:
        project_path = _BASE
    if not os.path.isdir(project_path):
        return {"error": f"ไม่พบโฟลเดอร์: {project_path}"}, 400

    _current_user["username"] = username
    _api_stop = {"flag": False}
    summary, static_findings, ai_report = security_review.run_security_review(
        project_path,
        should_stop=lambda: _api_stop["flag"]
    )
    return {"summary": summary, "static_findings": static_findings, "ai_report": ai_report}


# ── สร้างเอกสาร: Word / Excel / PowerPoint ──

@app.route("/api/doc/word", methods=["POST"])
def gen_word():
    """สร้างเอกสาร Word (.docx) — AI ร่างเนื้อหา → python-docx"""
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    if not _has_credit(username):
        return {"error": "เครดิตหมดแล้ว กรุณาเติมเงิน"}, 402
    data = request.get_json(force=True) or {}
    topic = (data.get("topic") or "").strip()
    hint = (data.get("hint") or "").strip()
    if not topic:
        return {"error": "ต้องระบุหัวข้อ"}, 400
    _current_user["username"] = username
    try:
        filepath, note = docgen.generate_word(topic, hint, username=username, llm=llm)
        if filepath:
            rel = os.path.relpath(filepath, _BASE)
            return {"ok": True, "file": rel, "note": note}
        return {"error": note}, 500
    except Exception as e:
        return {"error": str(e)[:300]}, 500


@app.route("/api/doc/excel", methods=["POST"])
def gen_excel():
    """สร้างตาราง Excel (.xlsx) — AI จัดข้อมูล → openpyxl"""
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    if not _has_credit(username):
        return {"error": "เครดิตหมดแล้ว กรุณาเติมเงิน"}, 402
    data = request.get_json(force=True) or {}
    topic = (data.get("topic") or "").strip()
    hint = (data.get("hint") or "").strip()
    if not topic:
        return {"error": "ต้องระบุหัวข้อ"}, 400
    _current_user["username"] = username
    try:
        filepath, note = docgen.generate_excel(topic, hint, username=username, llm=llm)
        if filepath:
            rel = os.path.relpath(filepath, _BASE)
            return {"ok": True, "file": rel, "note": note}
        return {"error": note}, 500
    except Exception as e:
        return {"error": str(e)[:300]}, 500


@app.route("/api/doc/pptx", methods=["POST"])
def gen_pptx():
    """สร้าง PowerPoint (.pptx) — AI สร้างสไลด์ → python-pptx"""
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    if not _has_credit(username):
        return {"error": "เครดิตหมดแล้ว กรุณาเติมเงิน"}, 402
    data = request.get_json(force=True) or {}
    topic = (data.get("topic") or "").strip()
    hint = (data.get("hint") or "").strip()
    if not topic:
        return {"error": "ต้องระบุหัวข้อ"}, 400
    _current_user["username"] = username
    try:
        filepath, note = docgen.generate_pptx(topic, hint, username=username, llm=llm)
        if filepath:
            rel = os.path.relpath(filepath, _BASE)
            return {"ok": True, "file": rel, "note": note}
        return {"error": note}, 500
    except Exception as e:
        return {"error": str(e)[:300]}, 500


@app.route("/api/sessions")
def sessions():
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    return {"sessions": _list_sessions(username=username)}


# ── Profile API ──

@app.route("/api/profile", methods=["GET"])
def get_profile():
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    ppath = _profile_path(username)
    if os.path.exists(ppath):
        try:
            with open(ppath, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}  # ยังไม่มี profile

@app.route("/api/profile", methods=["POST"])
def save_profile():
    login_check = require_login()
    if login_check:
        return login_check
    username = session.get("username")
    data = request.get_json(force=True) or {}
    career_group = (data.get("career_group") or "").strip()
    if not career_group:
        career_group = "ข้ามการตั้งค่า"
    profile = {
        "career_group": career_group,
        "skills": data.get("skills") or [],
        "style": data.get("style") or [],
        "project": data.get("project") or {},
        "updated_at": datetime.now().isoformat()
    }
    ppath = _profile_path(username)
    try:
        with open(ppath, "w", encoding="utf-8") as f:
            json.dump(profile, f, ensure_ascii=False, indent=2)
    except Exception as e:
        return {"error": f"บันทึกไม่สำเร็จ: {str(e)[:200]}"}, 500
    return {"ok": True}


def _user_cost_file(username):
    """Cost tracking file per user"""
    cost_dir = os.path.join(_BASE, "costs")
    os.makedirs(cost_dir, exist_ok=True)
    return os.path.join(cost_dir, f"{username or 'default'}.json")


_USD_TO_THB = 37.0
_DEFAULT_CREDIT_THB = 50.0


def _load_user_wallet(username):
    """Load user wallet: {total_usd, credit_thb, ...}"""
    try:
        with open(_user_cost_file(username), encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"total_usd": 0.0, "credit_thb": _DEFAULT_CREDIT_THB if username != "admin" else 999999.0}


def _save_user_wallet(username, data):
    try:
        with open(_user_cost_file(username), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception:
        pass


def _get_user_cost(username):
    """Get user's total cost in USD"""
    return _load_user_wallet(username).get("total_usd", 0.0)


def _get_user_credit_thb(username):
    """Get user's remaining credit in THB"""
    if username == "admin":
        return 999999.0
    return _load_user_wallet(username).get("credit_thb", _DEFAULT_CREDIT_THB)


def _has_credit(username):
    """Check if user has remaining credit"""
    if username == "admin":
        return True
    return _get_user_credit_thb(username) > 0


def _add_user_cost(username, tier, usage=None):
    """Add cost for a tier call — คิดตาม token จริง + peak/valley, หักเครดิต THB"""
    if usage and usage.get("cost_usd", 0) > 0:
        cost_usd = usage["cost_usd"]
    else:
        cost_usd = llm._config().get("cost_estimate_per_call_usd", {}).get(tier, 0.0)
    cost_thb = cost_usd * _USD_TO_THB
    w = _load_user_wallet(username)
    w["total_usd"] = w.get("total_usd", 0.0) + cost_usd
    if username != "admin":
        w["credit_thb"] = max(0.0, w.get("credit_thb", _DEFAULT_CREDIT_THB) - cost_thb)
    w["last_tier"] = tier
    w["last_cost_usd"] = cost_usd
    w["last_peak"] = usage.get("peak", False) if usage else None
    w["last_tokens_in"] = usage.get("input_tokens", 0) if usage else 0
    w["last_tokens_out"] = usage.get("output_tokens", 0) if usage else 0
    w["last_update"] = datetime.now().isoformat()
    _save_user_wallet(username, w)
    return w["total_usd"]


@app.route("/api/system")
def system():
    """System information for admin monitoring"""
    login_check = require_login()
    if login_check:
        return login_check
    
    username = session.get("username")
    
    # Get system info
    import psutil
    import platform
    
    # Count active users (sessions)
    active_users = 0
    try:
        session_files = glob.glob(os.path.join(app.config["SESSION_FILE_DIR"], "*"))
        # Check recent sessions (last 30 minutes)
        now = datetime.now()
        for f in session_files:
            try:
                if os.path.getmtime(f) > (now.timestamp() - 1800):  # 30 minutes
                    active_users += 1
            except:
                continue
    except:
        active_users = 1  # Fallback
    
    user_cost_usd = _get_user_cost(username)
    user_cost_thb = round(user_cost_usd * _USD_TO_THB, 2)
    user_wallet = _load_user_wallet(username)
    is_peak = llm._is_peak_hour()
    
    return {
        "display_name": session.get("display_name", ""),
        "username": username,
        "is_admin": username == "admin",
        "has_profile": _has_real_profile(username),
        "system": {
            "platform": platform.system(),
            "python": platform.python_version(),
            "cpu_percent": psutil.cpu_percent(interval=1),
            "memory_percent": psutil.virtual_memory().percent,
            "disk_percent": psutil.disk_usage('C:\\').percent
        },
        "application": {
            "active_users": active_users,
            "available_tiers": llm.available_tiers(),
            "total_chat_files": len(glob.glob(os.path.join(_CHAT_DIR, "**", "*.json"), recursive=True)),
            "total_sessions": len(glob.glob(os.path.join(_LOG_DIR, "session_*.json"))),
            "user_cost_thb": user_cost_thb,
            "user_cost_usd": round(user_cost_usd, 4),
            "credit_thb": round(_get_user_credit_thb(username), 2),
            "budget_today_usd": round(llm._today_cost(), 4),
            "is_peak_hour": is_peak,
            "peak_note": "⚠️ ช่วง Peak (2x ราคา)" if is_peak else "✓ ช่วง Valley (ราคาปกติ)",
            "error_lessons": error_memory.get_stats(),
            "last_call": {
                "tier": user_wallet.get("last_tier"),
                "cost_usd": user_wallet.get("last_cost_usd"),
                "peak": user_wallet.get("last_peak"),
                "tokens_in": user_wallet.get("last_tokens_in", 0),
                "tokens_out": user_wallet.get("last_tokens_out", 0),
            } if user_wallet.get("last_tier") else None
        }
    }


@app.route("/api/backups")
def backups():
    login_check = require_login()
    if login_check:
        return login_check
    return {"backups": applier.list_backups()}


@app.route("/api/restore", methods=["POST"])
def restore():
    login_check = require_login()
    if login_check:
        return login_check
    data = request.get_json(force=True) or {}
    bid = (data.get("id") or "").strip()
    if not bid:
        return {"error": "ไม่ได้ระบุแบ็กอัพที่จะย้อนคืน"}, 400
    try:
        r = applier.restore(bid)
    except applier.GuardError as e:
        return {"error": str(e)}, 400
    except Exception as e:
        return {"error": str(e)[:200]}, 500
    return {"ok": True, "result": r}


@app.route("/api/history")
def history():
    login_check = require_login()
    if login_check:
        return login_check
    who = (request.args.get("persona") or "").strip()
    if who not in personas.PERSONAS:
        return {"messages": []}
    username = session.get("username")
    return {"messages": _load_chat(who, username)}


@app.route("/api/inject", methods=["POST"])
def inject():
    """ฝากข้อความเข้าห้องแชทของน้องคนหนึ่ง (เช่น จดหมายจากพี่จูน) — ไม่เรียก AI"""
    login_check = require_login()
    if login_check:
        return login_check
    data = request.get_json(force=True) or {}
    who = (data.get("persona") or "").strip()
    role = (data.get("role") or "พี่จูน").strip()
    text = (data.get("text") or "").strip()
    if who not in personas.PERSONAS or not text:
        return {"error": "ต้องมี persona ที่ถูกต้อง + ข้อความ"}, 400
    username = session.get("username")
    msgs = _load_chat(who, username)
    msgs.append({"role": role, "text": text, "tier": data.get("tier")})
    _save_chat(who, msgs, username)
    return {"ok": True, "count": len(msgs)}


# ── พี่จูน (Claude) สถาปนิกอาวุโส วิเคราะห์งานของน้อง → วางพิมพ์เขียว ──
_BP_OPEN, _BP_CLOSE = "===พิมพ์เขียวสำหรับสภา===", "===จบพิมพ์เขียว==="
_SENIOR_SYSTEM = (
    "คุณคือ 'พี่จูน' (Claude) สถาปนิกซอฟต์แวร์อาวุโสของ Council Lab\n"
    "ผู้ใช้ส่งงานที่ทีม (แจ่มจูน/แจงจูน/เจนจูน) สรุป/ออกแบบไว้ มาให้คุณวิเคราะห์เชิงลึกและแนะนำเพิ่ม\n"
    "พูดอย่างมืออาชีพ ตรงไปตรงมา ใช้ศัพท์วิชาการ แทนตัวเองว่า 'พี่จูน' "
    "เรียกผู้ใช้ว่า 'คุณ' (หรือ 'อาจารย์' ถ้าผู้ใช้เป็น admin) "
    "ลงท้ายประโยคด้วย 'ค่ะ' ประโยคเด็ดขาด น่าเชื่อถือ ไม่ขี้อ้อน\n\n"
    "ทำ 4 ส่วนนี้ให้ครบ:\n"
    "1) 👍 จุดแข็งของแผนทีม — ชมตรงจุด\n"
    "2) ⚠️ จุดเสี่ยง/สิ่งที่ขาด — บอกตรงๆ พร้อมเหตุผล (เช่น edge case, ความปลอดภัย, scalability)\n"
    "3) 💡 คำแนะนำเพิ่มเติม — ทำให้ดีขึ้นยังไง เป็นรูปธรรม\n"
    "4) พิมพ์เขียวสั่งสภาทำต่อ — เขียนสเปกชัดเจนที่สภา (แจ่มจูน/แจงจูน/เจนจูน) เอาไปลงมือได้ทันที\n"
    f"   ครอบด้วยเครื่องหมายนี้เป๊ะๆ:\n   {_BP_OPEN}\n   <สเปกงานสั้นกระชับ สั่งได้เลย>\n   {_BP_CLOSE}\n"
    "ปิดท้าย (หลัง ===จบพิมพ์เขียว===) ด้วยบรรทัดคำแนะนำชัดเจนว่าผู้ใช้ควรกดปุ่มไหนต่อ:\n"
    "  • กดปุ่ม 🏛️ ส่งให้สภาแก้ไข (ทบทวนก่อน) — ส่งพิมพ์เขียวไปกล่องสภา ผู้ใช้ทบทวน/แก้ไขได้ก่อนกดประชุม\n"
    "  ถ้ามีจุดที่ผู้ใช้ควรระวัง/ตัดสินใจก่อน ให้บอกเหตุผลสั้นๆ ด้วย"
)


@app.route("/api/senior", methods=["POST"])
def senior():
    """พี่จูน(Claude) วิเคราะห์ + วางพิมพ์เขียว — ใช้ได้ทั้งจากห้องน้อง (persona)
    หรือจากการประชุมโดยตรง (session_id) อย่างใดอย่างหนึ่งก็ได้"""
    login_check = require_login()
    if login_check:
        return login_check
    data = request.get_json(force=True) or {}
    who = (data.get("persona") or "").strip()
    sid = (data.get("session_id") or "").strip()
    if "claude" not in llm.available_tiers():
        return {"error": "พี่จูน (Claude) ยังไม่พร้อม — ตรวจ ANTHROPIC_API_KEY ใน .env"}, 400

    expansion = ""
    username = session.get("username")
    if who in personas.PERSONAS:
        msgs = _load_chat(who, username)
        expansion = next((m.get("text") or "" for m in reversed(msgs)
                          if m.get("role") == who), "")
    sess = (_session_brief(sid) + "\n\n") if sid else ""   # โหมดประหยัด: เฉพาะโค้ด+เทส+มติ
    if not sess and not expansion:
        return {"error": "ไม่มีข้อมูลให้วิเคราะห์ — เลือกการประชุม หรือให้น้องขยายความก่อน"}, 400

    title = "อาจารย์" if username == "admin" else "คุณ"
    mat = f"[สิ่งที่ {who} สรุป/ออกแบบให้{title}]\n{expansion[:8000]}\n\n" if expansion else ""
    user = (f"{sess}{mat}ช่วยวิเคราะห์เชิงลึกและวางพิมพ์เขียวให้ตามหน้าที่ของพี่จูนค่ะ")
    senior_sys = _SENIOR_SYSTEM.replace("คุณ' (หรือ 'อาจารย์' ถ้าผู้ใช้เป็น admin)", f"{title}'")
    _current_user["username"] = username
    try:
        reply, tier, _ = llm.call_tier("claude", senior_sys, user, max_tokens=8000)
    except Exception as e:
        return {"error": str(e)[:300]}, 500

    bp = ""
    if _BP_OPEN in reply:                         # ทนแม้คำตอบโดนตัดท้าย (ไม่มีเครื่องหมายปิด)
        bp = reply.split(_BP_OPEN, 1)[1]
        bp = bp.split(_BP_CLOSE, 1)[0].strip() if _BP_CLOSE in bp else bp.strip()

    if who in personas.PERSONAS:                   # เก็บเข้าห้องน้องเฉพาะเมื่อเรียกจากห้อง
        msgs.append({"role": "พี่จูน", "text": reply, "tier": tier})
        _save_chat(who, msgs, username)
    return {"reply": reply, "tier": tier, "blueprint": bp}


@app.route("/api/expand_session", methods=["POST"])
def expand_session():
    """ขยายความการประชุมโดยตรง (แจ่มจูน deepseek) — ไม่บันทึกลงห้องน้อง"""
    login_check = require_login()
    if login_check:
        return login_check
    data = request.get_json(force=True) or {}
    sid = (data.get("session_id") or "").strip()
    blk = _session_block(sid)
    if not blk:
        return {"error": "ไม่พบการประชุมนี้"}, 400
    p = personas.PERSONAS["แจ่มจูน"]
    _username = session.get("username")
    _title = "อาจารย์" if _username == "admin" else "คุณ"
    prompt = (blk + f"\n\nช่วยสรุปขยายความการประชุมนี้ให้{_title}อ่านเข้าใจง่ายขึ้น — เล่าเป็นขั้นตอน "
              "ว่าทำอะไร ตัดสินใจอะไร เจอปัญหาอะไรแก้ยังไง ผลลัพธ์เป็นยังไง และมีอะไรที่ผู้ใช้ควรรู้เพิ่ม")
    _current_user["username"] = _username
    try:
        reply, tier, _ = llm.call_tier(p["model"], p["system"] + _build_chat_wrap(session.get("username")), prompt, max_tokens=6000)
    except Exception as e:
        return {"error": str(e)[:300]}, 500
    return {"reply": reply, "tier": tier}


@app.route("/api/stream")
def stream():
    login_check = require_login()
    if login_check:
        return login_check
    def gen():
        while True:
            ev = _events.get()
            yield f"data: {json.dumps(ev, ensure_ascii=False)}\n\n"
            if ev.get("speaker") == "__done__":
                break
    return Response(gen(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


if __name__ == "__main__":
    print("=" * 56)
    print(" 🏛️  COUNCIL TRINITY LAB — Web Console")
    print(" เปิดเบราว์เซอร์ที่:  http://127.0.0.1:8091")
    print(" tier ที่ใช้ได้:", ", ".join(llm.available_tiers()) or "ไม่มี!")
    print("=" * 56)
    try:
        from waitress import serve
        print(" 🚀 Production server: Waitress (4 threads)")
        serve(app, host="0.0.0.0", port=8091, threads=4)
    except ImportError:
        print(" ⚠️ Waitress ไม่ได้ติดตั้ง — ใช้ Flask dev server (ไม่แนะนำ)")
        app.run(host="0.0.0.0", port=8091, threaded=True)
